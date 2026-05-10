"""Shared pytest helpers for synthetic-fixture construction."""

from __future__ import annotations

import csv
from pathlib import Path

import openpyxl


def make_csv_fixture(
    path: Path,
    *,
    logger_name: str = "z6-19570",
    config: int = 1,
    rows: list[tuple] = (),
) -> Path:
    """Write a minimal METER-style CSV with a sensible 3-row header.

    The header advertises Port 1 as a TEROS 12, Port 2 as ATMOS 14,
    Port 3 as ECRN-100, Port 7 as Battery, Port 8 as Barometer. ``rows``
    is a list of tuples whose first element is the timestamp string in
    DD/MM/YYYY HH:MM:SS form; subsequent columns are floats matching the
    advertised header.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    port_row = [logger_name,
                "Port1", "Port1", "Port1",
                "Port2", "Port2", "Port2",
                "Port3", "Port3",
                "Port7", "Port7",
                "Port8", "Port8"]
    type_row = [f"# Records: {len(rows)}",
                "TEROS 12 Moisture/Temp/EC",
                "TEROS 12 Moisture/Temp/EC",
                "TEROS 12 Moisture/Temp/EC",
                "ATMOS 14 Humidity/Temp/Barometer",
                "ATMOS 14 Humidity/Temp/Barometer",
                "ATMOS 14 Humidity/Temp/Barometer",
                "ECRN-100 Precipitation",
                "ECRN-100 Precipitation",
                "Battery", "Battery",
                "Barometer", "Barometer"]
    var_row = ["Timestamps",
               " m3/m3 Water Content",
               " degree_C Soil Temperature",
               " mS/cm Saturation Extract EC",
               " degree_C Air Temperature",
               " kPa Vapor Pressure",
               " kPa Atmospheric Pressure",
               " mm Precipitation",
               " mm/h Max Precip Rate",
               "% Battery Percent",
               " mV Battery Voltage",
               " kPa Reference Pressure",
               " degree_C Logger Temperature"]
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(port_row)
        w.writerow(type_row)
        w.writerow(var_row)
        for r in rows:
            w.writerow(r)
    return path


def make_xlsx_fixture(
    path: Path,
    *,
    logger_name: str = "z6-19570",
    sheets: dict[str, list[tuple]] | None = None,
) -> Path:
    """Write a minimal METER-style Excel snapshot with one or more configs.

    ``sheets`` maps a config label (e.g. ``"Config 1"``) to its rows. Each
    sheet's header advertises the same single TEROS 12 (Port 1), Battery
    (Port 7), Barometer (Port 8) layout — small but representative.
    """
    sheets = sheets or {"Config 1": []}
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for label, rows in sheets.items():
        ws = wb.create_sheet(f"Processed Data {label}")
        ws.append([logger_name, "Port 1", "Port 1", "Port 1",
                    "Port 7", "Port 7", "Port 8", "Port 8"])
        ws.append([f"Records: {len(rows)}",
                    "TEROS 12", "TEROS 12", "TEROS 12",
                    "Battery", "Battery", "Barometer", "Barometer"])
        ws.append(["Timestamp",
                    "m³/m³ Water Content",
                    "°C Soil Temperature",
                    "mS/cm Bulk EC",
                    "% Battery Percent",
                    "mV Battery Voltage",
                    "kPa Reference Pressure",
                    "°C Logger Temperature"])
        for r in rows:
            ws.append(list(r))
    wb.save(path)
    wb.close()
    return path


def make_metadata_fixture(path: Path) -> Path:
    """Write a Metadata.xlsx with just enough to satisfy parse_metadata.

    Replicates the awkward stacked-tables layout of the real Serials sheet
    in miniature: one TEROS 12 sensor in each of Sw / Con / For locations,
    plus a Dataloggers subtable referencing them.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Serials")

    ws.append(["Field ID", None, "SN", "Type", "Aux or FishTail",
                "Location", "Depth cm (middle pin)", "Location notes",
                "Logger", None, None, None])
    ws.append(["T12-00148668", 148668, "SMS1", "TEROS 12", "Aux",
                "SwB", 10, None, "Log2", None, None, None])
    ws.append(["T12-00148662", 148662, "SMS2", "TEROS 12", "Aux",
                "SwB", 40, None, "Log2", None, None, None])
    ws.append(["T12-00148645", 148645, "SMS11", "TEROS 12", "Aux",
                "ConG", None, None, "Log3", None, None, None])
    ws.append(["T12-00149371", 149371, "SMS17", "TEROS 12", "Fishtail",
                "ForJ", None, None, None, None, None, None])

    # Spacer rows + Dataloggers header.
    for _ in range(3):
        ws.append([None] * 12)
    ws.append(["Dataloggers", "No", "Name", "Type", "Password",
                "Location", "Port 1", "Port 2", "Port 3", "Port 4",
                "Port 5", "Port 6"])
    ws.append(["Z6-19570", 19570, "Log2", "Z6", "p", "Top of Swale",
                "SMS1", "SMS2", None, None, None, None])
    ws.append(["Z6-05511", 55110, "Log3", "Z6", "p", "Control slope",
                "SMS11", None, None, None, None, None])

    wb.save(path)
    wb.close()
    return path
