"""File readers for METER ZL6 logger exports.

Both readers (`read_logger_csv`, `read_logger_xlsx`) return the same long-format
polars DataFrame:

    timestamp        Datetime (us, naive — local time at field site)
    port             UInt8     (1-8)
    sensor_type      Utf8      ('TEROS12' | 'ATMOS14' | 'ECRN100' | 'BATTERY' | 'BAROMETER')
    variable         Utf8      canonical name from schema.VARIABLE_MAP
    value            Float64
    source_format    Utf8      'csv' or 'xlsx'
    source_file      Utf8      basename of the file the row came from
    config_label     Utf8      'Config 1', 'Config 5', ... (sheet/file label)

The headers in both formats use three rows: row 1 = 'Port N' labels, row 2 =
sensor type, row 3 = unit + variable name. Each data column is a single
(port, sensor_type, variable) triple. Multiple columns share a port.
"""

from __future__ import annotations

import csv
import re
from datetime import datetime
from pathlib import Path

import openpyxl
import polars as pl

from swale.schema import (
    KEPT_SENSOR_TYPES,
    normalize_sensor_type,
    normalize_variable,
    parse_port_label,
)

# CSV timestamps: '25/05/2024 01:15:00'.
_CSV_TS_FORMAT = "%d/%m/%Y %H:%M:%S"

# CSV files for a configuration are named '... -Configuration N-...csv'. We
# pull out the 'Configuration N' label so different configs from one logger
# can be tracked in the source_file metadata.
_CSV_CONFIG_RE = re.compile(r"-(Configuration\s*\d+)-", re.IGNORECASE)

# The 'Records: N' or '# Records: N' string in the second row's first cell.
_RECORDS_RE = re.compile(r"^#?\s*Records\s*:", re.IGNORECASE)


def _build_column_specs(
    port_row: list,
    type_row: list,
    var_row: list,
) -> list[tuple[int, int, str, str] | None]:
    """Combine the three header rows into per-column descriptors.

    Returns a list aligned with the data columns. Element 0 corresponds to
    the timestamp column (always None). Subsequent elements are either a
    tuple ``(col_idx, port, sensor_type, variable)`` or None when the column
    should be ignored (unknown sensor type, unparseable header, etc.).
    """
    n = max(len(port_row), len(type_row), len(var_row))
    specs: list[tuple[int, int, str, str] | None] = [None] * n
    for i in range(1, n):
        port = parse_port_label(port_row[i] if i < len(port_row) else None)
        st = normalize_sensor_type(type_row[i] if i < len(type_row) else None)
        var = normalize_variable(var_row[i] if i < len(var_row) else None)
        if port is None or st not in KEPT_SENSOR_TYPES or var is None:
            continue
        specs[i] = (i, port, st, var)
    return specs


def _frame_from_records(
    timestamps: list[datetime],
    values_by_col: dict[int, list[float | None]],
    specs: list[tuple[int, int, str, str] | None],
    source_format: str,
    source_file: str,
    config_label: str,
) -> pl.DataFrame:
    """Materialise per-column value lists into a tidy long DataFrame."""
    if not timestamps:
        return _empty_long_frame()
    frames: list[pl.DataFrame] = []
    for spec in specs:
        if spec is None:
            continue
        col_idx, port, sensor_type, variable = spec
        vals = values_by_col.get(col_idx, [])
        # Pad if a row was short for some reason (shouldn't happen on
        # well-formed exports, but guards against ragged rows).
        if len(vals) < len(timestamps):
            vals = vals + [None] * (len(timestamps) - len(vals))
        frames.append(pl.DataFrame({
            "timestamp":     timestamps,
            "port":          [port] * len(timestamps),
            "sensor_type":   [sensor_type] * len(timestamps),
            "variable":      [variable] * len(timestamps),
            "value":         vals,
            "source_format": [source_format] * len(timestamps),
            "source_file":   [source_file] * len(timestamps),
            "config_label":  [config_label] * len(timestamps),
        }, schema=_LONG_SCHEMA))
    if not frames:
        return _empty_long_frame()
    return pl.concat(frames, how="vertical")


_LONG_SCHEMA: dict[str, pl.DataType] = {
    "timestamp":     pl.Datetime("us"),
    "port":          pl.UInt8,
    "sensor_type":   pl.Utf8,
    "variable":      pl.Utf8,
    "value":         pl.Float64,
    "source_format": pl.Utf8,
    "source_file":   pl.Utf8,
    "config_label":  pl.Utf8,
}


def _empty_long_frame() -> pl.DataFrame:
    return pl.DataFrame(schema=_LONG_SCHEMA)


def _to_float(v: object) -> float | None:
    if v is None or v == "" or v == "Not Set":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


# ---------------------------------------------------------------------------
# CSV reader
# ---------------------------------------------------------------------------

def read_logger_csv(path: Path) -> pl.DataFrame:
    """Read one METER ZL6 'Configuration N' CSV into long format.

    Args:
        path: Path to a *processed* configuration CSV. Files matching
            'Raw-Configuration' should not be passed here; they hold raw
            integer counts in a different schema.

    Returns:
        Long DataFrame with the schema documented at module level. Empty
        DataFrame if the file has only a header.

    Raises:
        FileNotFoundError: If ``path`` doesn't exist.

    Examples:
        >>> df = read_logger_csv(Path("z6-19570 - top-Configuration 1.csv"))
        >>> df.schema["timestamp"]
        Datetime(time_unit='us', time_zone=None)

    See Also:
        read_logger_xlsx: Counterpart for the periodic Excel snapshots.
    """
    if not path.exists():
        raise FileNotFoundError(path)

    config_match = _CSV_CONFIG_RE.search(path.name)
    config_label = config_match.group(1) if config_match else "Config ?"

    # Three header rows are read with the csv module; the data body is read
    # with polars for speed. We rely on polars only for ingestion of the
    # numeric body — header parsing is straightforward enough in stdlib.
    with path.open(newline="", encoding="utf-8-sig") as fh:
        rdr = csv.reader(fh)
        try:
            port_row = next(rdr)
            type_row = next(rdr)
            var_row = next(rdr)
        except StopIteration:
            return _empty_long_frame()

    specs = _build_column_specs(port_row, type_row, var_row)
    if not any(s is not None for s in specs):
        return _empty_long_frame()

    # Body. Use polars for speed; force every column to Utf8 (avoids the
    # parser silently picking different dtypes across files when one column
    # is empty in a particular file). We then cast inside the per-column
    # extraction loop.
    body = pl.read_csv(
        path,
        has_header=False,
        skip_rows=3,
        encoding="utf8",
        infer_schema_length=0,
        schema_overrides=None,
        truncate_ragged_lines=True,
    )
    if body.height == 0:
        return _empty_long_frame()

    # Polars assigns columns 'column_1', 'column_2', ... starting at 1.
    cols = body.columns

    # Parse timestamps (column 0 = 'column_1').
    ts_series = body[cols[0]].str.strptime(
        pl.Datetime("us"), format=_CSV_TS_FORMAT, strict=False
    )
    keep_mask = ts_series.is_not_null()
    body = body.filter(keep_mask)
    if body.height == 0:
        return _empty_long_frame()
    timestamps = body[cols[0]].str.strptime(
        pl.Datetime("us"), format=_CSV_TS_FORMAT, strict=False
    ).to_list()

    # Per-column values dict, keyed by spec col_idx.
    values_by_col: dict[int, list[float | None]] = {}
    for spec in specs:
        if spec is None:
            continue
        col_idx, *_ = spec
        if col_idx >= len(cols):
            continue
        col_name = cols[col_idx]
        # Cast to Float64; non-numeric becomes null.
        vals = body[col_name].cast(pl.Float64, strict=False).to_list()
        values_by_col[col_idx] = vals

    return _frame_from_records(
        timestamps,
        values_by_col,
        specs,
        source_format="csv",
        source_file=path.name,
        config_label=config_label,
    )


# ---------------------------------------------------------------------------
# Excel reader
# ---------------------------------------------------------------------------

# Sheet name pattern: 'Processed Data Config 5'. We collect every such sheet
# in a workbook; ignore Raw and Metadata sheets.
_PROCESSED_SHEET_RE = re.compile(r"^Processed\s+Data\s+(Config(?:uration)?\s*\d+)\s*$",
                                  re.IGNORECASE)


def read_logger_xlsx(path: Path) -> pl.DataFrame:
    """Read every 'Processed Data Config N' sheet of an Excel snapshot.

    A single Excel snapshot can hold multiple configurations stacked as
    sheets, each with its own three-row header. We iterate every matching
    sheet and concatenate the per-sheet long frames.

    Args:
        path: Path to a *.xlsx export from the METER cloud / Decagon app.

    Returns:
        Long DataFrame combining every processed-data sheet in the file.

    Raises:
        FileNotFoundError: If ``path`` doesn't exist.

    Examples:
        >>> df = read_logger_xlsx(Path("z6-19570 02Aug24-1459.xlsx"))
        >>> sorted(df["config_label"].unique().to_list())
        ['Config 1']

    See Also:
        read_logger_csv: Counterpart for the per-logger directory CSVs.
    """
    if not path.exists():
        raise FileNotFoundError(path)

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    frames: list[pl.DataFrame] = []
    try:
        for sheet_name in wb.sheetnames:
            m = _PROCESSED_SHEET_RE.match(sheet_name)
            if not m:
                continue
            config_label = m.group(1)
            ws = wb[sheet_name]
            it = ws.iter_rows(values_only=True)
            try:
                port_row = list(next(it))
                type_row = list(next(it))
                var_row = list(next(it))
            except StopIteration:
                continue
            specs = _build_column_specs(port_row, type_row, var_row)
            if not any(s is not None for s in specs):
                continue

            # Iterate body and collect into per-column lists. We materialise
            # timestamps once and value-lists per kept-column. Excel stores
            # timestamps as datetime objects already; just guard against the
            # occasional string export.
            timestamps: list[datetime] = []
            values_by_col: dict[int, list[float | None]] = {
                spec[0]: [] for spec in specs if spec is not None
            }
            for row in it:
                if not row:
                    continue
                ts_cell = row[0]
                ts: datetime | None
                if isinstance(ts_cell, datetime):
                    ts = ts_cell
                elif isinstance(ts_cell, str) and ts_cell.strip():
                    try:
                        ts = datetime.strptime(ts_cell.strip(), _CSV_TS_FORMAT)
                    except ValueError:
                        continue
                else:
                    continue
                timestamps.append(ts)
                for col_idx in values_by_col:
                    cell = row[col_idx] if col_idx < len(row) else None
                    values_by_col[col_idx].append(_to_float(cell))

            frames.append(_frame_from_records(
                timestamps,
                values_by_col,
                specs,
                source_format="xlsx",
                source_file=path.name,
                config_label=config_label,
            ))
    finally:
        wb.close()

    if not frames:
        return _empty_long_frame()
    return pl.concat(frames, how="vertical")
