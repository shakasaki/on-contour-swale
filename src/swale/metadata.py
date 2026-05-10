"""Metadata parser for the swale project.

Reads the 'Serials' tab of Metadata.xlsx and exposes two polars DataFrames:

* sensors:      one row per non-Forest sensor, with location, depth, treatment.
* port_mapping: one row per (logger_serial, port) pair, mapping to sensor_id.

The Excel layout is unusual: a single tab holds two stacked tables.

  Rows 1-25  : sensor inventory (Field ID, SN, Type, Aux/Fishtail,
               Location, Depth, notes, Logger).
  Row  31    : a header for a 'Dataloggers' subtable.
  Rows 32-34 : one row per logger with port-by-port sensor assignments.

Forest sensors (Location starts with 'For') are dropped, as is logger 05533.
"""

from __future__ import annotations

import re
import warnings
from pathlib import Path

import openpyxl
import polars as pl

from swale.schema import FOREST_LOGGER, LOGGER_INFO, NON_FOREST_LOGGERS

# Locations starting with these prefixes belong to each treatment. Anything
# else (e.g. 'Step??', '?') gets treatment=null and is excluded from analysis.
TREATMENT_PREFIXES: dict[str, str] = {
    "Sw":  "swale",
    "Con": "control",
    "For": "forest",
}


def _treatment_from_location(loc: str | None) -> str | None:
    if not isinstance(loc, str):
        return None
    for prefix, treatment in TREATMENT_PREFIXES.items():
        if loc.startswith(prefix):
            return treatment
    return None


def _normalize_sms_id(raw: object) -> str | None:
    """Map values like 'SMS1', 'SMS 16', 'SMS16' to a canonical 'SMS01' form.

    Inconsistent spacing in the source spreadsheet is normalized so we can
    match port-mapping entries (e.g. 'SMS16' in the Dataloggers subtable)
    against sensor inventory entries (e.g. 'SMS 16' in the Serials list).
    Returns None for non-string or unparseable inputs.
    """
    if not isinstance(raw, str):
        return None
    m = re.match(r"^\s*SMS\s*0*(\d+)\s*$", raw, re.IGNORECASE)
    if not m:
        return None
    return f"SMS{int(m.group(1)):02d}"


def _coerce_serial(raw: object) -> str | None:
    """Loggers serials and TEROS-12 SNs are sometimes ints, sometimes strings."""
    if raw is None:
        return None
    if isinstance(raw, float):
        return str(int(raw))
    return str(raw).strip() or None


def _coerce_depth(raw: object) -> int | None:
    if raw is None or raw == "":
        return None
    try:
        return int(float(raw))
    except (TypeError, ValueError):
        return None


def parse_metadata(metadata_xlsx: Path) -> tuple[pl.DataFrame, pl.DataFrame]:
    """Parse the Serials tab of the project metadata spreadsheet.

    Args:
        metadata_xlsx: Path to Metadata.xlsx.

    Returns:
        A tuple ``(sensors, port_mapping)``:

        * ``sensors`` columns: ``sensor_id`` (e.g. 'SMS01'), ``field_id``
          (e.g. 'T12-00148668'), ``serial_number`` (e.g. '148668'),
          ``sensor_type`` (e.g. 'TEROS12'), ``location`` ('SwB', 'ConG'…),
          ``treatment`` ('swale', 'control'), ``depth_cm`` (int or null),
          ``location_notes``.
        * ``port_mapping`` columns: ``logger_serial`` (5-digit, e.g.
          '19574'), ``port`` (1-8), ``sensor_id`` ('SMS01'…). Synthesised
          rows are added for the logger-internal Battery and Barometer
          channels so the loader can tag them uniformly.

    Raises:
        FileNotFoundError: If ``metadata_xlsx`` doesn't exist.
        ValueError: If the 'Serials' tab is missing, or the Dataloggers
            subtable cannot be located.

    Examples:
        >>> sensors, ports = parse_metadata(Path("Metadata.xlsx"))
        >>> sensors.filter(pl.col("treatment") == "swale").height
        9

    See Also:
        swale.loader.load_swale_dataset: Top-level ingestion entry point.
    """
    if not metadata_xlsx.exists():
        raise FileNotFoundError(metadata_xlsx)

    wb = openpyxl.load_workbook(metadata_xlsx, data_only=True, read_only=True)
    if "Serials" not in wb.sheetnames:
        raise ValueError("Metadata.xlsx is missing the 'Serials' sheet")
    ws = wb["Serials"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    sensor_records: list[dict] = []
    dataloggers_header_row: int | None = None

    for i, row in enumerate(rows):
        # Sensor inventory header (as of May 2026):
        #   0: Field ID, 1: <blank>, 2: SN, 3: Type, 4: Aux/Fishtail,
        #   5: Location, 6: tag, 7: treatment, 8: Depth cm, 9: Location notes,
        #   10: Logger
        field_id, _, sms_code, type_label, *rest = (row + (None,) * 12)[:12]
        # rest[0]=Aux/Fishtail, rest[1]=Location, rest[2]=tag,
        # rest[3]=treatment, rest[4]=Depth, rest[5]=Notes
        location = rest[1]
        tag = rest[2]
        explicit_treatment = rest[3]
        depth = rest[4]
        notes = rest[5]

        # Detect the Dataloggers header row to slice the subtable below.
        if isinstance(field_id, str) and field_id.strip().lower() == "dataloggers":
            dataloggers_header_row = i
            continue

        sid = _normalize_sms_id(sms_code)
        if sid is None:
            continue
        if not isinstance(type_label, str) or "TEROS" not in type_label.upper():
            continue

        # Treatment: prefer the explicit column when present and recognised;
        # otherwise fall back to the location-prefix derivation. We deliberately
        # stay strict about prefixes so that 'Step??' (SMS10) and '?'
        # (SMS23/SMS24) keep treatment=null until the user resolves them.
        derived = _treatment_from_location(location)
        if isinstance(explicit_treatment, str):
            t = explicit_treatment.strip().lower()
            treatment = t if t in {"swale", "control", "forest"} else derived
        else:
            treatment = derived
        if treatment == "forest":
            continue

        sensor_records.append({
            "sensor_id":      sid,
            "field_id":       field_id if isinstance(field_id, str) else None,
            "serial_number":  _coerce_serial(row[1]),
            "sensor_type":    "TEROS12",
            "location":       location if isinstance(location, str) else None,
            "tag":            tag.strip() if isinstance(tag, str) else None,
            "treatment":      treatment,
            "depth_cm":       _coerce_depth(depth),
            "location_notes": notes if isinstance(notes, str) else None,
        })

    if dataloggers_header_row is None:
        raise ValueError("Could not locate the 'Dataloggers' subtable in Serials")

    # Subtable header (current layout): index 0..7 are bookkeeping (the
    # 'tag' / 'treatment' columns inserted in the inventory section above
    # leave 6 and 7 blank here), and Port 1..6 sit at indices 8..13.
    # We discover the Port-1 column from the header row so we don't have to
    # rebreak the parser the next time someone tweaks columns.
    header_row = rows[dataloggers_header_row]
    try:
        port1_col = next(
            i for i, v in enumerate(header_row)
            if isinstance(v, str) and v.strip().lower() == "port 1"
        )
    except StopIteration:
        raise ValueError("Could not find 'Port 1' column in Dataloggers header")

    port_records: list[dict] = []
    for row in rows[dataloggers_header_row + 1:]:
        if not row:
            continue
        # Logger-row marker is a string like 'Z6-19574' in column 0.
        first = row[0]
        if not isinstance(first, str) or not first.upper().startswith("Z6-"):
            continue
        serial = first.split("-", 1)[1].strip()
        # Pad 5-digit serials so the metadata key matches '19574', '05511', etc.
        # (One spreadsheet entry has SN=55110 — that's a typo for 05511.)
        serial = serial.zfill(5)
        if serial == FOREST_LOGGER:
            continue
        if serial not in NON_FOREST_LOGGERS:
            warnings.warn(f"Unknown logger serial {serial!r} in Dataloggers table")
            continue
        for port_num, cell in enumerate(row[port1_col:port1_col + 6], start=1):
            sid = _normalize_sms_id(cell) if cell else None
            if sid is None:
                continue
            port_records.append({
                "logger_serial": serial,
                "port":          port_num,
                "sensor_id":     sid,
                "sensor_type":   "TEROS12",
            })

    # Synthesise rows for the logger-internal Battery (port 7) and Barometer
    # (port 8) channels so the loader can join uniformly. These channels
    # always live at those two ports across every export we have.
    for serial in NON_FOREST_LOGGERS:
        port_records.append({
            "logger_serial": serial,
            "port":          7,
            "sensor_id":     f"BATT_{serial}",
            "sensor_type":   "BATTERY",
        })
        port_records.append({
            "logger_serial": serial,
            "port":          8,
            "sensor_id":     f"BARO_{serial}",
            "sensor_type":   "BAROMETER",
        })
    # Synthesised rows for the ATMOS 14 (port 5) and ECRN-100 (port 6) on
    # logger 19570. Other loggers don't have these external sensors.
    port_records.append({
        "logger_serial": "19570", "port": 5,
        "sensor_id": "ATMOS14_19570", "sensor_type": "ATMOS14",
    })
    port_records.append({
        "logger_serial": "19570", "port": 6,
        "sensor_id": "ECRN100_19570", "sensor_type": "ECRN100",
    })

    sensors = pl.DataFrame(sensor_records, schema={
        "sensor_id":      pl.Utf8,
        "field_id":       pl.Utf8,
        "serial_number":  pl.Utf8,
        "sensor_type":    pl.Utf8,
        "location":       pl.Utf8,
        "tag":            pl.Utf8,
        "treatment":      pl.Utf8,
        "depth_cm":       pl.Int16,
        "location_notes": pl.Utf8,
    })
    port_mapping = pl.DataFrame(port_records, schema={
        "logger_serial": pl.Utf8,
        "port":          pl.UInt8,
        "sensor_id":     pl.Utf8,
        "sensor_type":   pl.Utf8,
    })

    # Sensors with missing depth or unmapped location are kept (per spec)
    # but reported once at parse time so they're visible.
    missing_depth = sensors.filter(pl.col("depth_cm").is_null())
    if missing_depth.height:
        warnings.warn(
            f"{missing_depth.height} sensors have no depth_cm in metadata: "
            f"{missing_depth['sensor_id'].to_list()}"
        )
    unknown_treatment = sensors.filter(pl.col("treatment").is_null())
    if unknown_treatment.height:
        warnings.warn(
            f"{unknown_treatment.height} sensors have an unrecognized location: "
            f"{unknown_treatment['sensor_id'].to_list()}"
        )

    # Decorate port_mapping with logger alias / location for downstream joins.
    port_mapping = port_mapping.with_columns([
        pl.col("logger_serial").map_elements(
            lambda s: LOGGER_INFO.get(s, {}).get("alias"),
            return_dtype=pl.Utf8,
        ).alias("logger_alias"),
        pl.col("logger_serial").map_elements(
            lambda s: LOGGER_INFO.get(s, {}).get("location"),
            return_dtype=pl.Utf8,
        ).alias("logger_location"),
    ])

    return sensors, port_mapping
